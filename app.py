import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import io
import pandas as pd
import re

# 预定义专业数据（注意需要包含映射后的全称）
PROGRAM_DATA = {
    "24建筑电气与智能化": 127,
    "24建筑电气与智能化（对口）": 77,
    "24机械电子工程": 139,
    "24数据科学与大数据技术": 149,
    "24电气工程及其自动化": 160,

    "24汽车服务工程": 93,
    "24电子信息工程": 159,
    "24电子信息工程（对口）": 76,
    "24人工智能": 81,
    "24机械设计制造及其自动化": 159,  # 注意这里的全称与映射表一致
    "24物联网工程": 147
}

# 专业简称到全称的映射表（新增）
professional_mapping = {
     "物联网": "24物联网工程",
    "机械": "24机械设计制造及其自动化",
    "机械设计": "24机械设计制造及其自动化",
    "大数据": "24数据科学与大数据技术",
    "建筑电气": "24建筑电气与智能化",
    "建筑电气对口": "24建筑电气与智能化（对口）",
    "建筑电气（对口）": "24建筑电气与智能化（对口）",
    "机械电子": "24机械电子工程",
    "电气": "24电气工程及其自动化",
    "电气工程": "24电气工程及其自动化",
    "汽服": "24汽车服务工程",
    "电子": "24电子信息工程",
    "电子信息": "24电子信息工程",
    "电子信息对口": "24电子信息工程（对口）",
    "电子信息（对口）": "24电子信息工程（对口）",
    "人工": "24人工智能",
    "人工智能": "24人工智能",

    # 可根据需要扩展更多简称
}


def get_week_columns(day):
    """获取星期对应的缺勤列和出勤率列"""
    day_column_map = {
        "星期一": ("E", "F"),
        "星期二": ("H", "I"),
        "星期三": ("K", "L"),
        "星期四": ("N", "O"),
        "星期日": ("Q", "R")
    }
    return day_column_map.get(day, ("", ""))  # 默认返回空列


def parse_input(text):
    pattern = re.compile(
        r'\d{2}:\d{2}:\d{2}\s*'  # 匹配时间戳（如 19:37:17）及后续空格
        r'([\s\S]*?)(?=\n\s*\d{2}:\d{2}:\d{2}|\Z)'  # 非贪婪匹配内容，直到下一个时间戳或文本结束
    )

    # 提取所有匹配项并清理空白
    groups = [group.strip() for group in pattern.findall(text) if group.strip()]

    # 用分号连接每组结果
    result = ';'.join(groups)

    # 将结果赋值给 text 变量
    text = result

    # 在每个分组的第一个指定符号后添加括号
    symbols = ["：", ":", " "]
    groups = text.split(';')
    new_groups = []

    for group in groups:
        if not group.strip():
            continue
        index = len(group)
        for symbol in symbols:
            pos = group.find(symbol)
            if pos != -1 and pos < index:
                index = pos
        if index < len(group):
            new_group = group[:index + 1] + "（）" + group[index + 1:]
        else:
            new_group = group
        new_groups.append(new_group)

    text = ';'.join(new_groups)
    #去除班级前面的中文数字
    groups = text.split(';')

    chinese_to_arabic = {
        '一': '1', '二': '2', '三': '3', '四': '4'
    }

    result_groups = []
    for group in groups:
        found = False
        new_group = ""
        for i, char in enumerate(group):
            if not found and char in chinese_to_arabic:
                new_group += chinese_to_arabic[char]
                found = True
            else:
                new_group += char
        result_groups.append(new_group)

    text = ';'.join(result_groups)
    # 预处理：在括号后添加空格防止粘连
    text = re.sub(r'（[^）]*）', lambda x: x.group(0) + ' ', text)

    attendance_dict = {}

    for record in re.split(r";\s*", text.strip("; ")):
        # 解析专业信息
        major_match = re.search(
            r"(\d{2,4})?\s*"  # 年级
            r"([\u4e00-\u9fa5]+(?:（[\u4e00-\u9fa5]+）)?)"  # 专业简称
            r"\d*班?",  # 班级信息
            record
        )
        if not major_match:
            continue

        # 构建完整专业名称
        year = major_match.group(1) or ""
        short_major = major_match.group(2)
        full_major = professional_mapping.get(short_major, f"{year}{short_major}")

        # 提取并清洗学生名单
        students = []
        for block in re.findall(r"([^（）]+)（缺勤）", record):
            # 同时过滤班级前缀和人数说明
            cleaned = re.sub(
                r"^([\u4e00-\u9fa5]+\d*班?缺勤\d*人?)[:：]?",
                "",
                block
            )
            # 分割学生姓名
            names = re.split(r"[、，,\s]+", cleaned.strip())
            students.extend([name for name in names if name])

        if students:
            attendance_dict.setdefault(full_major, []).extend(students)

    # 转换为指定的数据结构格式
    return [
        (major, absentees)
        for major, absentees in attendance_dict.items()
    ]
def input_attendance():
    """出勤数据输入表单"""
    with st.form("attendance_form"):
        day = st.selectbox("选择星期", ["星期一", "星期二", "星期三", "星期四", "星期日"])
        input_text = st.text_area("输入数据（支持专业简称，格式：专业班/简称: 姓名1、姓名2；可多行分号分隔）",
                                  placeholder="例：机械2班：张三、李四；大数据：王五、赵六（缺勤）",
                                  height=100)

        if st.form_submit_button("保存数据"):
            programs_and_absentees = parse_input(input_text)
            if not programs_and_absentees:
                st.error("未识别到有效专业和名单！请检查格式和专业简称（支持映射表内简称）")
                return

            if "attendance_data" not in st.session_state:
                st.session_state.attendance_data = {}

            for program, absent_list in programs_and_absentees:
                # 检查专业是否在预定义数据中
                if program not in PROGRAM_DATA:
                    st.warning(f"专业 {program} 不在预定义列表中，可能导致报表生成错误")
                    continue

                key = f"{day}_{program}"
                st.session_state.attendance_data[key] = {
                    "absentees": absent_list,
                    "total": PROGRAM_DATA[program],
                    "day": day,
                    "program": program
                }
            st.success("数据已保存！")


# 以下代码（calculate_attendance、update_worksheet、create_template_with_data、main）保持不变，仅修改了parse_input和新增映射表

def calculate_attendance():
    """计算所有出勤率"""
    if "attendance_data" not in st.session_state:
        return {}

    results = {}
    for key, data in st.session_state.attendance_data.items():
        absent_count = len(data["absentees"])
        attendance_rate = round((data["total"] - absent_count) / data["total"] * 100, 2)
        results[key] = {
            "absent_count": absent_count,
            "attendance_rate": attendance_rate,
            "absentees": ", ".join(data["absentees"]),
            "program": data["program"],
            "day": data["day"]
        }
    return results


def update_worksheet(ws, data):
    """更新工作表数据"""
    # 定位专业行号（数据行从第4行开始，专业在C列）
    program_rows = {row[2].value: idx for idx, row in enumerate(ws.iter_rows(min_row=4, max_row=14), start=4)}

    for key, values in data.items():
        day, program = key.split("_", 1)
        if program not in program_rows:
            continue

        row_num = program_rows[program]
        absent_col, rate_col = get_week_columns(values["day"])  # 使用数据中的day确保一致

        # 更新缺勤人数和出勤率
        ws[f"{absent_col}{row_num}"] = values["absent_count"]
        ws[f"{rate_col}{row_num}"] = f"{values['attendance_rate']}%"

        # 更新缺勤名单（在缺勤名单区域查找对应专业行）
        for absent_row in range(19, 30):  # 缺勤名单从第19行开始
            if ws.cell(row=absent_row, column=3).value == program:  # C列是专业列
                ws.cell(row=absent_row, column=4, value=values["absentees"])
                break


def create_template_with_data():
    """生成包含数据的模板并填充考勤数据"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 设置精确列宽
    ws.column_dimensions['A'].width = 5.45  # 学院
    ws.column_dimensions['B'].width = 8.36  # 辅导员
    ws.column_dimensions['C'].width = 26.27  # 专业
    ws.column_dimensions['D'].width = 8.36  # 人数
    ws.column_dimensions['E'].width = 7.09  # 星期一-缺勤
    ws.column_dimensions['F'].width = 8.36  # 出勤率
    ws.column_dimensions['G'].width = 10.91  # 备注
    ws.column_dimensions['H'].width = 7.09  # 星期二-缺勤
    ws.column_dimensions['I'].width = 8.36  # 出勤率
    ws.column_dimensions['J'].width = 10.91  # 备注
    ws.column_dimensions['K'].width = 7.09  # 星期三-缺勤
    ws.column_dimensions['L'].width = 8.36  # 出勤率
    ws.column_dimensions['M'].width = 10.91  # 备注
    ws.column_dimensions['N'].width = 7.09  # 星期四-缺勤
    ws.column_dimensions['O'].width = 8.36  # 出勤率
    ws.column_dimensions['P'].width = 10.91  # 备注
    ws.column_dimensions['Q'].width = 7.09  # 星期日-缺勤
    ws.column_dimensions['R'].width = 8.36  # 出勤率
    ws.column_dimensions['S'].width = 10.91  # 备注
    ws.column_dimensions['T'].width = 10.91  # 统计人
    ws.column_dimensions['U'].width = 12.82  # 平均出勤率

    # 设置行高
    ws.row_dimensions[1].height = 36  # 第一行
    ws.row_dimensions[2].height = 53  # 第二行
    for row in range(3, 40):  # 第三行及以下
        ws.row_dimensions[row].height = 17

    # 设置宋体12号字体
    song_font = Font(name='宋体', size=12)
    song_bold_font = Font(name='宋体', size=12, bold=True)

    # 设置标题行（合并A1:U1）
    ws.merge_cells('A1:U1')
    title_cell = ws['A1']
    title_cell.value = '2024-2025第2学期第7周24级学生晚自习出勤检查表'
    title_cell.font = song_bold_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 第一行表头（合并A2:D3）
    ws.merge_cells('A2:A3')
    ws['A2'] = "学院"
    ws.merge_cells('B2:B3')
    ws['B2'] = "辅导员"
    ws.merge_cells('C2:C3')
    ws['C2'] = "专业"
    ws.merge_cells('D2:D3')
    ws['D2'] = "人数"

    # 设置第一行表头的其他部分
    days = ["星期一", "星期二", "星期三", "星期四", "星期日"]
    for i, day in enumerate(days):
        col = chr(69 + i * 3)  # E(69), H(72), K(75), N(78), Q(81)
        ws.merge_cells(f'{col}2:{chr(ord(col) + 2)}2')  # 合并3列
        ws[f'{col}2'] = day
        ws[f'{col}2'].font = song_font
        ws[f'{col}2'].alignment = Alignment(horizontal='center', vertical='center')

    # 设置统计人和平均出勤率（合并T2:T3和U2:U3）
    ws.merge_cells('T2:T3')
    ws['T2'] = "统计人"
    ws.merge_cells('U2:U3')
    ws['U2'] = "平均出勤率"

    # 设置第二行表头（缺勤、出勤率、备注）
    sub_headers = ["缺勤", "出勤率", "备注"]
    for i in range(5):  # 5个星期
        for j, header in enumerate(sub_headers):
            col = chr(69 + i * 3 + j)  # E,F,G; H,I,J; etc.
            cell = ws[f'{col}3']
            cell.value = header
            cell.font = song_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 添加数据行（与模板完全一致）
    data_rows = [
        ["机电学院", "曹支红", "24建筑电气与智能化", 127] + [""] * 17,
        ["", "", "24建筑电气与智能化（对口）", 77] + [""] * 17,
        ["", "戴辰阳", "24机械电子工程", 139] + [""] * 17,
        ["", "", "24数据科学与大数据技术", 149] + [""] * 17,
        ["", "赵鹤", "24电气工程及其自动化", 160] + [""] * 17,
        ["", "", "24汽车服务工程", 93] + [""] * 17,
        ["", "马招弟", "24电子信息工程", 159] + [""] * 17,
        ["", "", "24电子信息工程（对口）", 76] + [""] * 17,
        ["", "", "24人工智能", 81] + [""] * 17,
        ["", "王曼曼", "24机械设计制造及其自动化", 159] + [""] * 17,
        ["", "", "24物联网工程", 147] + [""] * 17,
        ["", "", "小计：", "=SUM(D4:D14)"] + [""] * 17,
        ["", "", "全校合计：", ""] + [""] * 17
    ]

    for row_idx, row_data in enumerate(data_rows, start=4):
        ws.append(row_data)
        for cell in ws[row_idx]:
            cell.font = song_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 合并"机电学院"单元格（A4:A16）
    ws.merge_cells('A4:A16')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')

    # 合并辅导员单元格
    ws.merge_cells('B4:B5')  # 曹支红
    ws.merge_cells('B6:B7')  # 戴辰阳
    ws.merge_cells('B8:B9')  # 赵鹤
    ws.merge_cells('B10:B12')  # 马招弟
    ws.merge_cells('B13:B15')  # 王曼曼

    # 添加缺勤名单标题（合并A17:U17）
    ws.append(["以下为该周晚自习缺勤名单"])
    ws.merge_cells(f'A{ws.max_row}:U{ws.max_row}')
    title_cell = ws[f'A{ws.max_row}']
    title_cell.font = song_bold_font
    title_cell.alignment = Alignment(horizontal='center')

    # 缺勤名单表头（合并D18:U18）
    ws.append(["学院", "辅导员", "专业", "缺勤人员名单及次数"])
    ws.merge_cells(f'D{ws.max_row}:U{ws.max_row}')

    for cell in ws[18]:
        cell.font = song_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 添加缺勤名单数据行
    absentees_rows = [
        ["", "曹支红", "24建筑电气与智能化", ""],
        ["", "", "24建筑电气与智能化（对口）", ""],
        ["", "戴辰阳", "24机械电子工程", ""],
        ["", "", "24数据科学与大数据技术", ""],
        ["", "赵鹤", "24电气工程及其自动化", ""],
        ["", "", "24汽车服务工程", ""],
        ["", "马招弟", "24电子信息工程", ""],
        ["", "", "24电子信息工程（对口）", ""],
        ["", "", "24人工智能", ""],
        ["", "王曼曼", "24机械设计制造及其自动化", ""],
        ["", "", "24物联网工程", ""]
    ]

    for row_data in absentees_rows:
        ws.append(row_data)
        # 合并学院和辅导员单元格（根据模板结构）
        if row_data[1]:  # 辅导员有值时合并B列
            start_row = ws.max_row
            # 查找连续相同辅导员的行数（简化逻辑，实际需根据模板结构）
            ws.merge_cells(f'B{start_row}:B{start_row}')  # 保持模板结构，实际已在数据行合并

    # 设置所有单元格边框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=21):
        for cell in row:
            cell.border = thin_border

    # 填充考勤数据
    if "attendance_data" in st.session_state:
        calculated_data = calculate_attendance()
        update_worksheet(ws, calculated_data)

    # 保存到字节流
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def main():
    st.set_page_config(page_title="智能晚自习考勤系统", layout="wide")
    st.title("智能晚自习考勤系统")

    # 数据输入部分
    st.header("出勤数据录入")
    input_attendance()

    # 数据显示部分
    st.header("已录入数据")
    if "attendance_data" in st.session_state:
        df = pd.DataFrame.from_dict(
            st.session_state.attendance_data,
            orient="index",
            columns=["day", "program", "absentees", "total"]
        )
        st.dataframe(df, use_container_width=True)

    # 生成下载按钮
    st.header("生成考勤表")
    if st.button("生成考勤表", type="primary"):
        if "attendance_data" not in st.session_state or not st.session_state.attendance_data:
            st.warning("请先录入至少一条考勤数据")
            return

        with st.spinner("正在生成表格..."):
            try:
                buffer = create_template_with_data()
                st.download_button(
                    label="下载考勤表",
                    data=buffer,
                    file_name="智能考勤表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            except Exception as e:
                st.error(f"生成表格时出现错误：{str(e)}", icon="🚨")


if __name__ == "__main__":
    main()
